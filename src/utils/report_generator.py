"""
report_generator.py
───────────────────
Utility functions for generating PDF (ReportLab) and Excel/CSV reports
from the campaign analytics DataFrame returned by load_campaign_data().
"""

from __future__ import annotations

import io
from datetime import datetime

import pandas as pd

from src.utils.campaigns import calculate_metrics, calculate_revenue

# ──────────────────────────────────────────────────────────────────────────────
# Helpers
# ──────────────────────────────────────────────────────────────────────────────

def _compute_kpis(df: pd.DataFrame) -> dict:
    """Return a dict of headline KPIs from the campaign DataFrame using central calculate_metrics."""
    metrics = calculate_metrics(df)
    clicks = int(df["clicks"].sum()) if "clicks" in df.columns else 0
    impr = int(df["impressions"].sum()) if "impressions" in df.columns else 0

    return dict(
        spend=metrics["totalSpend"],
        clicks=clicks,
        impressions=impr,
        signups=int(metrics["totalSignups"]),
        conversions=int(metrics["totalActivations"]),
        revenue=metrics["totalRevenue"],
        ctr=metrics["ctr"] * 100 if metrics["ctr"] <= 1.0 else metrics["ctr"],
        cvr=metrics["cvr"] * 100 if metrics["cvr"] <= 1.0 else metrics["cvr"],
        cpa=metrics["cpa"],
        cpl=metrics["cpl"],
        roas=metrics["roas"],
        roi=metrics["roi"],
        cpc=metrics["cpc"],
        aov=metrics["aov"],
    )


def _campaign_summary(df: pd.DataFrame) -> pd.DataFrame:
    """Aggregate per-campaign totals."""
    df_temp = df.copy()
    if "revenue" not in df_temp.columns:
        df_temp["revenue"] = calculate_revenue(df_temp)

    name_col = "campaign_name" if "campaign_name" in df_temp.columns else (
        "campaign" if "campaign" in df_temp.columns else "campaign_id"
    )
    
    grp_cols = [name_col]
    platform_col = "ad_platform" if "ad_platform" in df_temp.columns else (
        "platform" if "platform" in df_temp.columns else None
    )
    if platform_col:
        grp_cols.append(platform_col)

    agg: dict = {}
    metric_map = {
        "spend_usd": ["spend_usd", "spend"],
        "clicks": ["clicks"],
        "impressions": ["impressions"],
        "signups": ["signups"],
        "activations_7d": ["activations_7d", "conversions", "activations"],
        "revenue": ["revenue"],
    }

    for target, options in metric_map.items():
        for opt in options:
            if opt in df_temp.columns:
                agg[opt] = "sum"
                break

    summary = df_temp.groupby(grp_cols, as_index=False).agg(agg)

    # Standardize column names
    if "spend" in summary.columns and "spend_usd" not in summary.columns:
        summary.rename(columns={"spend": "spend_usd"}, inplace=True)
    if "activations" in summary.columns and "activations_7d" not in summary.columns:
        summary.rename(columns={"activations": "activations_7d"}, inplace=True)
    elif "conversions" in summary.columns and "activations_7d" not in summary.columns:
        summary.rename(columns={"conversions": "activations_7d"}, inplace=True)

    # CTR
    if "clicks" in summary.columns and "impressions" in summary.columns:
        summary["ctr_%"] = (summary["clicks"] / summary["impressions"] * 100).fillna(0.0).round(2)
    
    # ROAS
    if "revenue" in summary.columns and "spend_usd" in summary.columns:
        summary["roas"] = (summary["revenue"] / summary["spend_usd"]).fillna(0.0).round(2)
        
    sort_col = "spend_usd" if "spend_usd" in summary.columns else summary.columns[0]
    return summary.sort_values(sort_col, ascending=False).reset_index(drop=True)


# ──────────────────────────────────────────────────────────────────────────────
# PDF generators (ReportLab)
# ──────────────────────────────────────────────────────────────────────────────

def _hex(h: str):
    """Convert '#rrggbb' to reportlab Color."""
    from reportlab.lib.colors import HexColor
    return HexColor(h)


def _base_doc(buf: io.BytesIO, title: str):
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.units import cm
    from reportlab.platypus import SimpleDocTemplate
    return SimpleDocTemplate(
        buf,
        pagesize=A4,
        topMargin=1.8 * cm,
        bottomMargin=1.8 * cm,
        leftMargin=2 * cm,
        rightMargin=2 * cm,
        title=title,
        author="CampaignCanvas",
    )


def _styles():
    from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
    ss = getSampleStyleSheet()

    heading = ParagraphStyle(
        "CCHeading",
        parent=ss["Heading1"],
        fontName="Helvetica-Bold",
        fontSize=22,
        textColor=_hex("#0284c7"),
        spaceAfter=4,
        leading=26,
    )
    sub = ParagraphStyle(
        "CCSub",
        parent=ss["Normal"],
        fontName="Helvetica",
        fontSize=9,
        textColor=_hex("#475569"),
        spaceAfter=14,
    )
    label = ParagraphStyle(
        "CCLabel",
        parent=ss["Normal"],
        fontName="Helvetica-Bold",
        fontSize=8,
        textColor=_hex("#334155"),
        spaceBefore=12,
        spaceAfter=2,
        leading=10,
    )
    body = ParagraphStyle(
        "CCBody",
        parent=ss["Normal"],
        fontName="Helvetica",
        fontSize=9,
        textColor=_hex("#334155"),
        spaceAfter=6,
        leading=13,
    )
    section = ParagraphStyle(
        "CCSection",
        parent=ss["Normal"],
        fontName="Helvetica-Bold",
        fontSize=12,
        textColor=_hex("#0f172a"),
        spaceBefore=18,
        spaceAfter=6,
    )
    return dict(heading=heading, sub=sub, label=label, body=body, section=section)


def _kpi_table(kpis: dict, cols: int = 4):
    """Build a clean KPI grid table."""
    from reportlab.lib.styles import ParagraphStyle
    from reportlab.platypus import Paragraph, Table, TableStyle

    items = [
        ("Total Spend", f"${kpis['spend']:,.0f}"),
        ("Revenue", f"${kpis['revenue']:,.0f}"),
        ("ROAS", f"{kpis['roas']:.2f}x"),
        ("ROI", f"{kpis['roi']:.1f}%"),
        ("Impressions", f"{kpis['impressions']:,}"),
        ("Clicks", f"{kpis['clicks']:,}"),
        ("CTR", f"{kpis['ctr']:.2f}%"),
        ("CVR", f"{kpis['cvr']:.2f}%"),
        ("CPC", f"${kpis['cpc']:.2f}"),
        ("CPA", f"${kpis['cpa']:.0f}"),
        ("CPL", f"${kpis['cpl']:.0f}"),
        ("AOV", f"${kpis['aov']:.0f}"),
        ("Signups", f"{kpis['signups']:,}"),
        ("Conversions", f"{kpis['conversions']:,}"),
    ]

    while len(items) % cols:
        items.append(("", ""))

    label_style = ParagraphStyle(
        "kpi_label",
        fontName="Helvetica",
        fontSize=7,
        textColor=_hex("#94a3b8"),
        leading=9,
        spaceAfter=3,
    )
    value_style = ParagraphStyle(
        "kpi_value",
        fontName="Helvetica-Bold",
        fontSize=14,
        textColor=_hex("#f8fafc"),
        leading=17,
    )

    col_width = (17.6 / cols) * 72 / 2.54
    row_height = 52

    grid_rows = []
    for i in range(0, len(items), cols):
        row_cells = []
        for j in range(cols):
            lbl, val = items[i + j]
            inner = Table(
                [[Paragraph(lbl, label_style)], [Paragraph(val, value_style)]],
                colWidths=[col_width - 16],
                rowHeights=[12, 22],
            )
            inner.setStyle(
                TableStyle(
                    [
                        ("TOPPADDING", (0, 0), (-1, -1), 0),
                        ("BOTTOMPADDING", (0, 0), (-1, -1), 0),
                        ("LEFTPADDING", (0, 0), (-1, -1), 0),
                        ("RIGHTPADDING", (0, 0), (-1, -1), 0),
                    ]
                )
            )
            row_cells.append(inner)
        grid_rows.append(row_cells)

    t = Table(
        grid_rows,
        colWidths=[col_width] * cols,
        rowHeights=[row_height] * len(grid_rows),
    )

    bg_colors = [_hex("#0f172a"), _hex("#1e293b")]
    style_cmds = [
        ("BOX", (0, 0), (-1, -1), 0.5, _hex("#334155")),
        ("INNERGRID", (0, 0), (-1, -1), 0.5, _hex("#334155")),
        ("TOPPADDING", (0, 0), (-1, -1), 10),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 10),
        ("LEFTPADDING", (0, 0), (-1, -1), 8),
        ("RIGHTPADDING", (0, 0), (-1, -1), 8),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
    ]
    for ri, row in enumerate(grid_rows):
        style_cmds.append(("BACKGROUND", (0, ri), (-1, ri), bg_colors[ri % 2]))

    t.setStyle(TableStyle(style_cmds))
    return t


def _campaign_table(summary: pd.DataFrame):
    from reportlab.lib.styles import ParagraphStyle
    from reportlab.platypus import Paragraph, Table, TableStyle

    display_cols = []
    rename = {}

    name_col = "campaign_name" if "campaign_name" in summary.columns else (
        "campaign" if "campaign" in summary.columns else "campaign_id"
    )
    display_cols.append(name_col)
    rename[name_col] = "Campaign"

    platform_col = "ad_platform" if "ad_platform" in summary.columns else (
        "platform" if "platform" in summary.columns else None
    )
    if platform_col:
        display_cols.append(platform_col)
        rename[platform_col] = "Platform"

    for c, label in [
        ("spend_usd", "Spend ($)"),
        ("spend", "Spend ($)"),
        ("revenue", "Revenue ($)"),
        ("roas", "ROAS"),
        ("clicks", "Clicks"),
        ("ctr_%", "CTR (%)"),
    ]:
        if c in summary.columns and c not in display_cols:
            display_cols.append(c)
            rename[c] = label

    sub = summary[display_cols].rename(columns=rename)

    header_style = ParagraphStyle(
        "th", fontName="Helvetica-Bold", fontSize=7, textColor=_hex("#ffffff")
    )
    cell_style = ParagraphStyle(
        "td", fontName="Helvetica", fontSize=7, textColor=_hex("#f1f5f9")
    )

    usable_pt = 17.6 * 72 / 2.54
    col_count = len(sub.columns)
    col_widths = [usable_pt / col_count] * col_count

    data = [[Paragraph(str(c), header_style) for c in sub.columns]]
    for _, row in sub.iterrows():
        formatted = []
        for c, v in zip(sub.columns, row):
            if isinstance(v, float):
                formatted.append(f"{v:,.2f}")
            elif isinstance(v, int):
                formatted.append(f"{v:,}")
            else:
                formatted.append(str(v))
        data.append([Paragraph(f, cell_style) for f in formatted])

    t = Table(data, colWidths=col_widths, repeatRows=1)
    t.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), _hex("#0ea5e9")),
                ("BACKGROUND", (0, 1), (-1, -1), _hex("#0f172a")),
                (
                    "ROWBACKGROUNDS",
                    (0, 1),
                    (-1, -1),
                    [_hex("#0f172a"), _hex("#1e293b")],
                ),
                ("BOX", (0, 0), (-1, -1), 0.5, _hex("#334155")),
                ("INNERGRID", (0, 0), (-1, -1), 0.25, _hex("#1e293b")),
                ("TOPPADDING", (0, 0), (-1, -1), 4),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
                ("LEFTPADDING", (0, 0), (-1, -1), 6),
                ("RIGHTPADDING", (0, 0), (-1, -1), 6),
            ]
        )
    )
    return t


def _footer_canvas(canvas, doc):
    """Draw a coloured footer bar on every page."""
    from reportlab.lib.pagesizes import A4

    w, _ = A4
    canvas.saveState()
    canvas.setFillColor(_hex("#0ea5e9"))
    canvas.rect(0, 0, w, 18, fill=1, stroke=0)
    canvas.setFont("Helvetica", 7)
    canvas.setFillColor(_hex("#f8fafc"))
    canvas.drawString(20, 5, "CampaignCanvas  ·  Confidential")
    canvas.drawRightString(
        w - 20,
        5,
        f"Generated {datetime.now().strftime('%Y-%m-%d %H:%M')}  ·  Page {doc.page}",
    )
    canvas.restoreState()


# ──────────────────────────────────────────────────────────────────────────────
# Public API
# ──────────────────────────────────────────────────────────────────────────────

def generate_executive_pdf(df: pd.DataFrame) -> bytes:
    """One-page executive summary with top KPI table only."""
    from reportlab.platypus import HRFlowable, Paragraph, Spacer

    buf = io.BytesIO()
    doc = _base_doc(buf, "Executive Summary")
    s = _styles()
    kpis = _compute_kpis(df)

    camp_count = (
        df["campaign_id"].nunique()
        if "campaign_id" in df.columns
        else (df["campaign"].nunique() if "campaign" in df.columns else "—")
    )

    story = [
        Paragraph("CampaignCanvas", s["heading"]),
        Paragraph("Executive Summary Report", s["sub"]),
        HRFlowable(
            width="100%", thickness=0.75, color=_hex("#cbd5e1"), spaceAfter=10
        ),
        Paragraph("Headline KPIs", s["section"]),
        Spacer(1, 4),
        _kpi_table(kpis, cols=4),
        Spacer(1, 18),
        HRFlowable(width="100%", thickness=0.75, color=_hex("#cbd5e1")),
        Spacer(1, 8),
        Paragraph(
            f"Data covers <b>{len(df):,}</b> rows across "
            f"<b>{camp_count}</b> campaigns. "
            f"All monetary values in USD.",
            s["body"],
        ),
    ]
    doc.build(story, onFirstPage=_footer_canvas, onLaterPages=_footer_canvas)
    return buf.getvalue()


def generate_summary_pdf(df: pd.DataFrame) -> bytes:
    """KPIs + top 10 campaign breakdown."""
    from reportlab.platypus import HRFlowable, Paragraph, Spacer

    buf = io.BytesIO()
    doc = _base_doc(buf, "Summary Report")
    s = _styles()
    kpis = _compute_kpis(df)
    summary = _campaign_summary(df).head(10)

    story = [
        Paragraph("CampaignCanvas", s["heading"]),
        Paragraph("Summary Analytics Report", s["sub"]),
        HRFlowable(
            width="100%", thickness=0.75, color=_hex("#cbd5e1"), spaceAfter=10
        ),
        Paragraph("Headline KPIs", s["section"]),
        Spacer(1, 4),
        _kpi_table(kpis, cols=4),
        Spacer(1, 18),
        Paragraph("Top Campaigns", s["section"]),
        Spacer(1, 4),
        _campaign_table(summary),
        Spacer(1, 12),
        HRFlowable(width="100%", thickness=0.75, color=_hex("#cbd5e1")),
        Spacer(1, 6),
        Paragraph(
            f"Showing top {len(summary)} campaigns by spend. "
            f"Full dataset: <b>{len(df):,}</b> rows.",
            s["body"],
        ),
    ]
    doc.build(story, onFirstPage=_footer_canvas, onLaterPages=_footer_canvas)
    return buf.getvalue()


def generate_detailed_pdf(df: pd.DataFrame) -> bytes:
    """Full campaign table + KPIs across all campaigns."""
    from reportlab.platypus import HRFlowable, Paragraph, Spacer

    buf = io.BytesIO()
    doc = _base_doc(buf, "Detailed Report")
    s = _styles()
    kpis = _compute_kpis(df)
    summary = _campaign_summary(df)

    story = [
        Paragraph("CampaignCanvas", s["heading"]),
        Paragraph("Detailed Campaign Report", s["sub"]),
        HRFlowable(
            width="100%", thickness=0.75, color=_hex("#cbd5e1"), spaceAfter=10
        ),
        Paragraph("Headline KPIs", s["section"]),
        Spacer(1, 4),
        _kpi_table(kpis, cols=4),
        Spacer(1, 20),
        Paragraph("All Campaigns", s["section"]),
        Spacer(1, 4),
        _campaign_table(summary),
        Spacer(1, 16),
        Paragraph("Raw Data Sample (first 20 rows)", s["section"]),
        Spacer(1, 4),
        _campaign_table(df.head(20)),
        Spacer(1, 12),
        HRFlowable(width="100%", thickness=0.75, color=_hex("#cbd5e1")),
        Spacer(1, 6),
        Paragraph(
            f"Full dataset: <b>{len(df):,}</b> rows  ·  "
            f"Generated {datetime.now().strftime('%Y-%m-%d %H:%M UTC')}",
            s["body"],
        ),
    ]
    doc.build(story, onFirstPage=_footer_canvas, onLaterPages=_footer_canvas)
    return buf.getvalue()


def generate_csv(df: pd.DataFrame) -> bytes:
    """Raw campaign DataFrame as UTF-8 CSV."""
    return df.to_csv(index=False).encode("utf-8")


def generate_excel(df: pd.DataFrame) -> bytes:
    """Multi-sheet Excel workbook: Raw Data, Campaign Summary, KPI Overview."""
    kpis = _compute_kpis(df)
    summary = _campaign_summary(df)

    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        # Sheet 1: Raw Data
        df.to_excel(writer, sheet_name="Raw Data", index=False)
        _style_excel_sheet(writer.sheets["Raw Data"], df)

        # Sheet 2: Campaign Summary
        summary.to_excel(writer, sheet_name="Campaign Summary", index=False)
        _style_excel_sheet(writer.sheets["Campaign Summary"], summary)

        # Sheet 3: KPI Overview
        kpi_df = pd.DataFrame(
            [
                ("Total Spend (USD)", f"${kpis['spend']:,.2f}"),
                ("Total Revenue (USD)", f"${kpis['revenue']:,.2f}"),
                ("ROAS", f"{kpis['roas']:.2f}x"),
                ("ROI", f"{kpis['roi']:.2f}%"),
                ("Total Impressions", f"{kpis['impressions']:,}"),
                ("Total Clicks", f"{kpis['clicks']:,}"),
                ("CTR", f"{kpis['ctr']:.2f}%"),
                ("Conversion Rate", f"{kpis['cvr']:.2f}%"),
                ("CPC", f"${kpis['cpc']:.2f}"),
                ("CPA", f"${kpis['cpa']:.2f}"),
                ("CPL", f"${kpis['cpl']:.2f}"),
                ("AOV", f"${kpis['aov']:.2f}"),
                ("Total Signups", f"{kpis['signups']:,}"),
                ("Total Conversions", f"{kpis['conversions']:,}"),
            ],
            columns=["Metric", "Value"],
        )
        kpi_df.to_excel(writer, sheet_name="KPI Overview", index=False)
        _style_excel_sheet(writer.sheets["KPI Overview"], kpi_df)

    return buf.getvalue()


def _style_excel_sheet(ws, df: pd.DataFrame):
    """Apply minimal styling to an openpyxl sheet."""
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    header_fill = PatternFill("solid", fgColor="0EA5E9")
    alt_fill = PatternFill("solid", fgColor="1E293B")
    alt2_fill = PatternFill("solid", fgColor="0F172A")
    header_font = Font(name="Calibri", bold=True, color="FFFFFF", size=10)
    cell_font = Font(name="Calibri", color="CBD5E1", size=9)
    thin = Side(style="thin", color="334155")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for col_idx, col in enumerate(df.columns, start=1):
        col_letter = get_column_letter(col_idx)
        header_cell = ws[f"{col_letter}1"]
        header_cell.fill = header_fill
        header_cell.font = header_font
        header_cell.alignment = Alignment(horizontal="center", vertical="center")
        header_cell.border = border

        # Auto-width calculation
        max_len = max(
            len(str(col)),
            df[col].astype(str).str.len().max() if not df.empty else 0,
        )
        ws.column_dimensions[col_letter].width = min(max_len + 4, 40)

    for row_idx in range(2, ws.max_row + 1):
        fill = alt_fill if row_idx % 2 == 0 else alt2_fill
        for cell in ws[row_idx]:
            cell.fill = fill
            cell.font = cell_font
            cell.border = border
            cell.alignment = Alignment(vertical="center")

    ws.freeze_panes = "A2"
    ws.sheet_view.showGridLines = False